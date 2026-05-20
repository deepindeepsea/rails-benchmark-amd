#!/usr/bin/env python3
"""
AMD Performance Excel Report Generator
Matches the Netflix/PerfSpect benchmark profile format.

Each workload run produces a sheet with two columns (System A / System B)
of PerfSpect-style metrics, making it easy to compare Genoa vs Turin,
baseline vs optimized, or any two AMD system configurations.

Usage:
    # Single system (one column of metrics):
    python3 amd_perf_excel_report.py --workload "sleep 2" \
        --label "Genoa 96C 320W" --output amd_report.xlsx

    # Two systems side-by-side (run on each and merge):
    python3 amd_perf_excel_report.py \
        --workload "openssl speed -seconds 5 aes-256-cbc" \
        --label "Genoa" --output genoa.json --json-only

    python3 amd_perf_excel_report.py \
        --workload "openssl speed -seconds 5 aes-256-cbc" \
        --label "Turin" --output turin.json --json-only

    python3 amd_perf_excel_report.py --merge genoa.json turin.json \
        --output amd_comparison.xlsx

    # Shortcut: compare two workloads on the same system:
    python3 amd_perf_excel_report.py \
        --workload "openssl speed aes-256-cbc" --label "OpenSSL" \
        --workload2 "dd if=/dev/zero of=/dev/null bs=1M count=500" --label2 "DD" \
        --sheet-name "openssl_vs_dd" --output amd_report.xlsx
"""

import sys
import json
import subprocess
import platform
import datetime
import os
import argparse

# ── openpyxl ─────────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
except ImportError:
    print("Installing openpyxl…")
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl",
                    "--break-system-packages", "-q"], check=True)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference


# ─────────────────────────────────────────────────────────────────────────────
# Event definitions  (confirmed working on AMD Zen4/Zen5 baremetal)
# ─────────────────────────────────────────────────────────────────────────────

EVENT_GROUPS = {
    "pipeline_l1": (
        "de_no_dispatch_per_slot.no_ops_from_frontend,"
        "de_no_dispatch_per_slot.backend_stalls,"
        "de_src_op_disp.all,"
        "ex_ret_ops,"
        "ls_not_halted_cyc"
    ),
    "backend_breakdown": (
        "ex_no_retire.load_not_complete,"
        "ex_no_retire.not_complete,"
        "ls_not_halted_cyc"
    ),
    "branch_prediction": (
        "ex_ret_brn_misp,"
        "ex_ret_brn,"
        "cpu-cycles,"
        "instructions"
    ),
    "l2_cache": (
        "l2_cache_req_stat.dc_hit_in_l2,"
        "l2_cache_req_stat.ls_rd_blk_c,"
        "l2_cache_req_stat.ic_fill_miss,"
        "l2_cache_req_stat.ic_hit_in_l2"
    ),
}


# ─────────────────────────────────────────────────────────────────────────────
# perf stat collection
# ─────────────────────────────────────────────────────────────────────────────

def collect_events(events: str, workload: str, verbose=True) -> dict:
    cmd = f'perf stat -j -e "{events}" -- {workload} 2>&1'
    if verbose:
        print(f"  Running: perf stat -j -e \"{events[:60]}...\" -- {workload[:40]}")
    try:
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=180)
        output = result.stdout + result.stderr
    except subprocess.TimeoutExpired:
        print("  Warning: perf stat timed out", file=sys.stderr)
        return {}

    values = {}
    for line in output.splitlines():
        line = line.strip()
        if not line or '"event"' not in line:
            continue
        try:
            obj = json.loads(line)
            event = obj.get("event", "").strip()
            val_str = obj.get("counter-value", "0").replace(",", "").strip()
            if val_str in ("<not counted>", "<not supported>", ""):
                val = 0.0
            else:
                val = float(val_str)
            if event:
                values[event] = val
        except (json.JSONDecodeError, ValueError):
            pass
    return values


def safe_div(num, denom, default=0.0):
    try:
        if denom == 0:
            return default
        return num / denom
    except Exception:
        return default


# ─────────────────────────────────────────────────────────────────────────────
# Metric Calculation  (matches PerfSpect metric names)
# ─────────────────────────────────────────────────────────────────────────────

def calculate_metrics(events: dict, cpu_info: str) -> list:
    """
    Returns an ordered list of (metric_name, value) tuples matching the
    Netflix/PerfSpect sheet layout.  Unknown/unsupported metrics get None.
    """

    # Raw events
    frontend     = events.get("de_no_dispatch_per_slot.no_ops_from_frontend", 0)
    backend      = events.get("de_no_dispatch_per_slot.backend_stalls", 0)
    dispatched   = events.get("de_src_op_disp.all", 0)
    retired      = events.get("ex_ret_ops", 0)
    cycles_nhc   = events.get("ls_not_halted_cyc", 1)   # non-halted cycles

    load_nc      = events.get("ex_no_retire.load_not_complete", 0)
    not_complete = events.get("ex_no_retire.not_complete", 1)

    misp         = events.get("ex_ret_brn_misp", 0)
    branches     = events.get("ex_ret_brn", 1)
    cpu_cycles   = events.get("cpu-cycles", 1)
    instructions = events.get("instructions", 0)

    l2_dc_hits   = events.get("l2_cache_req_stat.dc_hit_in_l2", 0)
    l2_dc_miss   = events.get("l2_cache_req_stat.ls_rd_blk_c", 0)
    l2_ic_miss   = events.get("l2_cache_req_stat.ic_fill_miss", 0)
    l2_ic_hits   = events.get("l2_cache_req_stat.ic_hit_in_l2", 0)

    # Derived
    ipc   = safe_div(instructions, cpu_cycles)
    cpi   = safe_div(cpu_cycles, instructions)
    gipc  = safe_div(instructions, cpu_cycles * 1e9) if cpu_cycles > 0 else 0.0

    total_slots  = cycles_nhc * 6
    fe_pct   = safe_div(frontend, total_slots) * 100
    be_pct   = safe_div(backend, total_slots) * 100
    bs_pct   = safe_div(max(dispatched - retired, 0), total_slots) * 100
    ret_pct  = safe_div(retired, total_slots) * 100

    mem_ratio = safe_div(load_nc, not_complete)
    be_mem_pct = be_pct * mem_ratio
    be_cpu_pct = be_pct * (1.0 - mem_ratio)

    misp_ratio = safe_div(misp, branches)

    # L2 hits per thousand instructions (PTI) — matches PerfSpect naming
    k = 1000.0 / max(instructions, 1)
    l2_dc_hits_pti = l2_dc_hits * k
    l2_dc_miss_pti = l2_dc_miss * k
    l2_ic_hits_pti = l2_ic_hits * k
    l2_ic_miss_pti = l2_ic_miss * k

    l2_dc_hit_rate = safe_div(l2_dc_hits, l2_dc_hits + l2_dc_miss + 1e-9) * 100
    l2_ic_hit_rate = safe_div(l2_ic_hits, l2_ic_hits + l2_ic_miss + 1e-9) * 100

    metrics = [
        # ── Section: CPU Basics ──────────────────────────────────────────
        ("_SECTION_", "CPU Basics"),
        ("CPU info",                        cpu_info),
        ("CPI",                             round(cpi, 6)),
        ("IPC",                             round(ipc, 6)),
        ("giga_instructions_per_sec",       round(gipc * cpu_cycles / 1e9, 4)
                                            if cpu_cycles > 0 else None),

        # ── Section: Branch Prediction ──────────────────────────────────
        ("_SECTION_", "Branch Prediction"),
        ("Branch Misprediction Ratio",      round(misp_ratio, 6)),
        ("Branches Retired (M)",            round(branches / 1e6, 2)),
        ("Branch Mispredicts (M)",          round(misp / 1e6, 2)),

        # ── Section: L2 Cache (PTI = per thousand instructions) ─────────
        ("_SECTION_", "L2 Cache"),
        ("L2 Cache Hits from L1 Data Cache Misses PTI",
                                            round(l2_dc_hits_pti, 6)),
        ("L2 Cache Misses from L1 Data Cache Misses PTI",
                                            round(l2_dc_miss_pti, 6)),
        ("L2 Cache Hits from L1 Instruction Cache Misses PTI",
                                            round(l2_ic_hits_pti, 6)),
        ("L2 Cache Misses from L1 Instruction Cache Misses PTI",
                                            round(l2_ic_miss_pti, 6)),
        ("L2 Data Cache Hit Rate (%)",      round(l2_dc_hit_rate, 4)),
        ("L2 Instruction Cache Hit Rate (%)", round(l2_ic_hit_rate, 4)),

        # ── Section: Pipeline Utilization ───────────────────────────────
        ("_SECTION_", "Pipeline Utilization"),
        ("Pipeline Utilization - Frontend Bound (%)",       round(fe_pct, 6)),
        ("Pipeline Utilization - Bad Speculation (%)",      round(bs_pct, 6)),
        ("Pipeline Utilization - Backend Bound (%)",        round(be_pct, 6)),
        ("Pipeline Utilization - Backend Bound - Memory (%)", round(be_mem_pct, 6)),
        ("Pipeline Utilization - Backend Bound - CPU (%)",  round(be_cpu_pct, 6)),
        ("Pipeline Utilization - SMT Contention (%)",       0.0),
        ("Pipeline Utilization - Retiring (%)",             round(ret_pct, 6)),

        # ── Section: Raw Counters ───────────────────────────────────────
        ("_SECTION_", "Raw Counters"),
        ("Active CPU Cycles (M)",           round(cycles_nhc / 1e6, 2)),
        ("Total Dispatch Slots (M)",        round(total_slots / 1e6, 2)),
        ("Frontend Unused Slots (M)",       round(frontend / 1e6, 2)),
        ("Backend Unused Slots (M)",        round(backend / 1e6, 2)),
        ("Dispatched Ops (M)",              round(dispatched / 1e6, 2)),
        ("Retired Ops (M)",                 round(retired / 1e6, 2)),
        ("Instructions Retired (M)",        round(instructions / 1e6, 2)),
        ("Non-Retire Events (M)",           round(not_complete / 1e6, 2)),
        ("Load Not Complete Events (M)",    round(load_nc / 1e6, 2)),
    ]

    return metrics


# ─────────────────────────────────────────────────────────────────────────────
# Style helpers
# ─────────────────────────────────────────────────────────────────────────────

# Netflix-style colors
AMD_RED     = "FFE53935"   # AMD/header red
AMD_DARK    = "FF1A1A2E"   # dark header
SECTION_BG  = "FFF5F5F5"   # light grey section separator
GOOD_GREEN  = "FFD4EDDA"
WARN_AMBER  = "FFFFF3CD"
BAD_RED     = "FFF8D7DA"
WHITE       = "FFFFFFFF"
LIGHT_BLUE  = "FFE3F2FD"

def header_font(size=12):
    return Font(bold=True, color="FFFFFFFF", size=size)

def section_font():
    return Font(bold=True, color="FF1A1A2E", size=10, italic=True)

def value_font():
    return Font(size=10, name="Consolas")

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="FFCCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


# ─────────────────────────────────────────────────────────────────────────────
# Excel sheet builder
# ─────────────────────────────────────────────────────────────────────────────

PIPELINE_METRICS = {
    "Pipeline Utilization - Frontend Bound (%)":       ("low", 35),
    "Pipeline Utilization - Bad Speculation (%)":      ("low", 20),
    "Pipeline Utilization - Backend Bound (%)":        ("low", 45),
    "Pipeline Utilization - Backend Bound - Memory (%)": ("low", 25),
    "Pipeline Utilization - Backend Bound - CPU (%)":  ("low", 25),
    "Pipeline Utilization - Retiring (%)":             ("high", 40),
    "IPC":                                             ("high", 1.5),
    "Branch Misprediction Ratio":                      ("low", 0.10),
    "L2 Data Cache Hit Rate (%)":                      ("high", 80),
    "L2 Instruction Cache Hit Rate (%)":               ("high", 80),
}


def color_for(metric_name, value):
    """Return fill color based on good/bad thresholds."""
    if metric_name not in PIPELINE_METRICS or value is None:
        return None
    direction, threshold = PIPELINE_METRICS[metric_name]
    try:
        v = float(value)
    except (TypeError, ValueError):
        return None
    if direction == "low":
        if v < threshold * 0.6:
            return GOOD_GREEN
        if v > threshold:
            return BAD_RED
        return WARN_AMBER
    else:  # high — bigger is better
        if v > threshold:
            return GOOD_GREEN
        if v < threshold * 0.5:
            return BAD_RED
        return WARN_AMBER


def write_sheet(wb, sheet_name, metrics_a, label_a,
                metrics_b=None, label_b=None,
                workload_a="", workload_b="", timestamp=""):
    """
    Write one sheet in the Netflix PerfSpect format.
    metrics_a / metrics_b are lists of (name, value) from calculate_metrics().
    """
    ws = wb.create_sheet(title=sheet_name[:31])  # Excel 31-char sheet name limit

    # ── Column widths ──────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14

    # ── Title row ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = f"AMD Performance Profile — {sheet_name}"
    title_cell.font = Font(bold=True, color="FFFFFFFF", size=14)
    title_cell.fill = make_fill(AMD_DARK)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ── Sub-header row ─────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 16
    for col, val, align in [
        ("A", "Generated: " + timestamp,  "left"),
        ("B", "Workload A: " + workload_a[:30], "left"),
        ("C", "Workload B: " + (workload_b or "")[:30] if workload_b else "", "left"),
        ("D", "", "left"),
    ]:
        c = ws[col + "2"]
        c.value = val
        c.font = Font(italic=True, color="FF888888", size=9)
        c.alignment = Alignment(horizontal=align)

    # ── Column header row ──────────────────────────────────────────────────
    ws.row_dimensions[3].height = 22
    headers = [
        ("A3", "Metric"),
        ("B3", label_a),
        ("C3", label_b if label_b else ""),
        ("D3", "% Delta" if label_b else ""),
    ]
    for cell_ref, heading in headers:
        c = ws[cell_ref]
        c.value = heading
        c.font = header_font(10)
        c.fill = make_fill(AMD_RED)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

    # ── Convert metrics lists to dicts for lookup ──────────────────────────
    def metrics_to_dict(metrics):
        return {name: val for name, val in metrics if not name.startswith("_SECTION_")}

    dict_a = metrics_to_dict(metrics_a)
    dict_b = metrics_to_dict(metrics_b) if metrics_b else {}

    # ── Write metric rows following the order of metrics_a ─────────────────
    row = 4
    for name, val_a in metrics_a:
        ws.row_dimensions[row].height = 16

        if name == "_SECTION_":
            # Section divider row
            ws.merge_cells(f"A{row}:D{row}")
            c = ws[f"A{row}"]
            c.value = f"  {val_a}"
            c.font = section_font()
            c.fill = make_fill(SECTION_BG)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.border = thin_border()
            row += 1
            continue

        # Metric name
        c_name = ws[f"A{row}"]
        c_name.value = name
        c_name.font = Font(size=10)
        c_name.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        c_name.border = thin_border()

        # Value A
        c_a = ws[f"B{row}"]
        c_a.value = val_a
        c_a.font = value_font()
        c_a.alignment = Alignment(horizontal="right", vertical="center")
        c_a.border = thin_border()
        fill_a = color_for(name, val_a)
        if fill_a:
            c_a.fill = make_fill(fill_a)

        # Value B
        c_b = ws[f"C{row}"]
        val_b = dict_b.get(name)
        c_b.value = val_b
        c_b.font = value_font()
        c_b.alignment = Alignment(horizontal="right", vertical="center")
        c_b.border = thin_border()
        fill_b = color_for(name, val_b)
        if fill_b and label_b:
            c_b.fill = make_fill(fill_b)

        # % Delta
        c_d = ws[f"D{row}"]
        if label_b and val_b is not None and val_a is not None:
            try:
                delta = (float(val_b) / float(val_a) - 1) if float(val_a) != 0 else None
                if delta is not None:
                    c_d.value = delta
                    c_d.number_format = "0.0%"
                    c_d.font = Font(size=10,
                                    bold=True,
                                    color="FF2E7D32" if delta > 0.02 else
                                          ("FFC62828" if delta < -0.02 else "FF000000"))
            except (TypeError, ValueError):
                pass
        c_d.alignment = Alignment(horizontal="right", vertical="center")
        c_d.border = thin_border()

        row += 1

    # ── Pipeline chart ─────────────────────────────────────────────────────
    # Find the pipeline rows
    pipeline_names = [
        "Pipeline Utilization - Frontend Bound (%)",
        "Pipeline Utilization - Bad Speculation (%)",
        "Pipeline Utilization - Backend Bound - Memory (%)",
        "Pipeline Utilization - Backend Bound - CPU (%)",
        "Pipeline Utilization - SMT Contention (%)",
        "Pipeline Utilization - Retiring (%)",
    ]

    # Collect row numbers for chart data
    chart_rows_a = []
    chart_rows_b = []
    chart_labels_row = []
    data_start = 4
    for r in range(data_start, row):
        cell_val = ws[f"A{r}"].value
        if cell_val in pipeline_names:
            chart_labels_row.append(r)
            chart_rows_a.append(ws[f"B{r}"].value or 0)
            if label_b:
                chart_rows_b.append(ws[f"C{r}"].value or 0)

    if chart_labels_row:
        chart = BarChart()
        chart.type = "bar"
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.title = "Pipeline Slot Distribution (%)"
        chart.y_axis.title = "% of dispatch slots"
        chart.x_axis.title = "System"
        chart.width = 22
        chart.height = 14

        # Data refs
        labels_ref = Reference(ws,
                                min_col=1, min_row=chart_labels_row[0],
                                max_row=chart_labels_row[-1])
        data_ref_a = Reference(ws, min_col=2, min_row=chart_labels_row[0],
                               max_row=chart_labels_row[-1])
        chart.add_data(data_ref_a, titles_from_data=False)
        chart.series[0].title.v = label_a

        if label_b and chart_rows_b:
            data_ref_b = Reference(ws, min_col=3, min_row=chart_labels_row[0],
                                   max_row=chart_labels_row[-1])
            chart.add_data(data_ref_b, titles_from_data=False)
            chart.series[1].title.v = label_b

        chart.set_categories(labels_ref)

        chart_anchor = f"F4"
        ws.add_chart(chart, chart_anchor)

    return ws


# ─────────────────────────────────────────────────────────────────────────────
# Data collection wrapper
# ─────────────────────────────────────────────────────────────────────────────

def collect_all_metrics(workload: str, label: str) -> list:
    cpu_info = get_cpu_info()
    print(f"\n── Collecting AMD perf events for: {label} ──")
    print(f"   CPU:      {cpu_info}")
    print(f"   Workload: {workload}")

    all_events = {}
    for group_name, events in EVENT_GROUPS.items():
        print(f"   Group [{group_name}]...", end=" ", flush=True)
        group_data = collect_events(events, workload, verbose=False)
        all_events.update(group_data)
        active = sum(1 for v in group_data.values() if v > 0)
        print(f"({active}/{len(group_data)} events active)")

    metrics = calculate_metrics(all_events, cpu_info)
    return metrics


def get_cpu_info() -> str:
    try:
        result = subprocess.run(["lscpu"], capture_output=True, text=True, timeout=5)
        for line in result.stdout.splitlines():
            if "Model name" in line:
                return line.split(":", 1)[1].strip()
    except Exception:
        pass
    return platform.processor() or "Unknown CPU"


# ─────────────────────────────────────────────────────────────────────────────
# Summary sheet
# ─────────────────────────────────────────────────────────────────────────────

def write_summary_sheet(wb, sheet_names, timestamp):
    ws = wb.create_sheet(title="README", index=0)
    ws.column_dimensions["A"].width = 60

    ws["A1"].value = "AMD Performance Analysis Workbook"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFFFF")
    ws["A1"].fill = make_fill(AMD_DARK)
    ws.row_dimensions[1].height = 32

    ws["A2"].value = f"Generated: {timestamp}"
    ws["A2"].font = Font(italic=True, color="FF888888", size=10)

    ws["A4"].value = "Methodology"
    ws["A4"].font = Font(bold=True, size=12)
    notes = [
        "• Events collected via Linux perf stat (JSON output mode)",
        "• AMD-specific dispatch slot model: 6 slots per cycle",
        "• Pipeline categories: Frontend / Backend / Bad Speculation / Retiring",
        "• Backend subdivided: Memory-bound vs CPU-bound (via load_not_complete ratio)",
        "• L2 cache metrics expressed as PTI (per thousand instructions)",
        "• Color coding: Green = good, Amber = marginal, Red = needs attention",
        "",
        "Sheets in this workbook:",
    ]
    for i, note in enumerate(notes, start=5):
        ws[f"A{i}"].value = note
        ws[f"A{i}"].font = Font(size=10)

    for i, name in enumerate(sheet_names, start=len(notes) + 5):
        ws[f"A{i}"].value = f"  → {name}"
        ws[f"A{i}"].font = Font(size=10, color="FF1565C0", underline="single")

    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="AMD Performance Excel Report Generator (PerfSpect style)"
    )
    parser.add_argument("--workload",  default="sleep 2",
                        help='Workload A command (default: "sleep 2")')
    parser.add_argument("--label",    default="System A",
                        help="Label for System A column")
    parser.add_argument("--workload2", default=None,
                        help="Optional Workload B for side-by-side comparison")
    parser.add_argument("--label2",   default="System B",
                        help="Label for System B column")
    parser.add_argument("--sheet-name", default=None,
                        help="Sheet name (defaults to first word of workload)")
    parser.add_argument("--output",   default="amd_perf_report.xlsx",
                        help="Output Excel file (default: amd_perf_report.xlsx)")
    parser.add_argument("--json-only", action="store_true",
                        help="Save raw metrics as JSON (for later --merge)")
    parser.add_argument("--merge",    nargs=2, metavar=("A.json", "B.json"),
                        help="Merge two previously saved JSON metric files")
    args = parser.parse_args()

    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── Merge mode ─────────────────────────────────────────────────────────
    if args.merge:
        file_a, file_b = args.merge
        with open(file_a) as f:
            data_a = json.load(f)
        with open(file_b) as f:
            data_b = json.load(f)

        sheet_name = args.sheet_name or "comparison"
        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # remove default sheet
        write_summary_sheet(wb, [sheet_name], timestamp)
        write_sheet(wb, sheet_name,
                    metrics_a=[(k, v) for k, v in data_a["metrics"]],
                    label_a=data_a["label"],
                    metrics_b=[(k, v) for k, v in data_b["metrics"]],
                    label_b=data_b["label"],
                    workload_a=data_a["workload"],
                    workload_b=data_b["workload"],
                    timestamp=timestamp)
        out = args.output
        wb.save(out)
        print(f"\nExcel report saved to: {os.path.abspath(out)}")
        return

    # ── Single or dual workload mode ───────────────────────────────────────
    metrics_a = collect_all_metrics(args.workload, args.label)

    if args.json_only:
        out = args.output
        payload = {
            "label":   args.label,
            "workload": args.workload,
            "metrics": metrics_a,
            "timestamp": timestamp,
        }
        with open(out, "w") as f:
            json.dump(payload, f, indent=2)
        print(f"\nJSON metrics saved to: {os.path.abspath(out)}")
        return

    metrics_b = None
    label_b   = None
    workload_b = None
    if args.workload2:
        metrics_b  = collect_all_metrics(args.workload2, args.label2)
        label_b    = args.label2
        workload_b = args.workload2

    sheet_name = args.sheet_name or args.workload.split()[0].replace("/", "-")[:31]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_summary_sheet(wb, [sheet_name], timestamp)
    write_sheet(wb, sheet_name,
                metrics_a=metrics_a,
                label_a=args.label,
                metrics_b=metrics_b,
                label_b=label_b,
                workload_a=args.workload,
                workload_b=workload_b or "",
                timestamp=timestamp)

    out = args.output
    wb.save(out)
    print(f"\nExcel report saved to: {os.path.abspath(out)}")
    print("Open in Excel/LibreOffice Calc to view with conditional formatting.")

    # Print quick summary to terminal
    print("\nQuick Summary:")
    for name, val in metrics_a:
        if name == "_SECTION_":
            print(f"\n  [{val}]")
        elif val is not None:
            print(f"    {name:<55} {val}")


if __name__ == "__main__":
    main()
